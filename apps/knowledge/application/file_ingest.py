"""
File ingest: layered processing from raw file to extracted text and metadata.

Layers: raw file → extracted text + metadata → (downstream) sanitized text → indexed/stored chunks.

Status contract:
- ok: text extracted and above length threshold; safe for downstream PII/chunking.
- empty: no text; do not index.
- scanned_review_required: text too short (likely scanned/image PDF). OCR is NOT implemented:
  we do not run OCR on scanned PDFs; the document is marked for manual review and not
  auto-indexed. Do not assume scanned PDFs are fully supported.
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from typing import Any, BinaryIO, Optional

# Minimum non-whitespace chars to consider document as text-extracted (not scanned)
MIN_EXTRACTED_TEXT_LEN = 50

# Status after extraction
STATUS_OK = "ok"
STATUS_EMPTY = "empty"
STATUS_SCANNED_REVIEW_REQUIRED = "scanned_review_required"


@dataclass
class FileMetadata:
    """Document metadata from the file (may contain PII; handle with care)."""

    filename: str = ""
    author: Optional[str] = None
    creator: Optional[str] = None
    modified_by: Optional[str] = None
    modified: Optional[str] = None  # ISO date string if available
    raw: dict[str, Any] = field(default_factory=dict)


@dataclass
class ExtractedFileResult:
    """Result of the extraction layer: text + metadata + status."""

    extracted_text: str
    metadata: FileMetadata
    status: str  # ok | empty | scanned_review_required
    is_scanned_review_required: bool = False
    mime_type: str = ""


def _normalize(text: str) -> str:
    return (text or "").strip()


def _is_sparse(text: str) -> bool:
    """True if text is empty or too short (likely scanned image PDF)."""
    n = len(_normalize(text))
    return n < MIN_EXTRACTED_TEXT_LEN


def extract_pdf(stream: BinaryIO, filename: str) -> ExtractedFileResult:
    """Extract text and metadata from PDF. Uses pdfplumber if available, else pdfminer."""
    metadata = FileMetadata(filename=filename or "document.pdf")
    text_parts: list[str] = []

    try:
        import pdfplumber
        with pdfplumber.open(stream) as pdf:
            if getattr(pdf, "metadata", None):
                meta = pdf.metadata
                metadata.author = meta.get("Author") or meta.get("/Author")
                metadata.creator = meta.get("Creator") or meta.get("/Creator")
                metadata.raw = dict(meta)
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    text_parts.append(t)
    except ImportError:
        from pdfminer.high_level import extract_text_to_fp
        from io import StringIO
        stream.seek(0)
        out = StringIO()
        extract_text_to_fp(stream, out)
        text_parts.append(out.getvalue())
    except Exception:
        text_parts = []

    extracted = "\n".join(text_parts)
    normalized = _normalize(extracted)
    if not normalized:
        return ExtractedFileResult(
            extracted_text="",
            metadata=metadata,
            status=STATUS_EMPTY,
            is_scanned_review_required=True,
            mime_type="application/pdf",
        )
    if _is_sparse(normalized):
        return ExtractedFileResult(
            extracted_text=extracted,
            metadata=metadata,
            status=STATUS_SCANNED_REVIEW_REQUIRED,
            is_scanned_review_required=True,
            mime_type="application/pdf",
        )
    return ExtractedFileResult(
        extracted_text=extracted,
        metadata=metadata,
        status=STATUS_OK,
        is_scanned_review_required=False,
        mime_type="application/pdf",
    )


def extract_docx(stream: BinaryIO, filename: str) -> ExtractedFileResult:
    """Extract text and metadata from DOCX."""
    from docx import Document

    metadata = FileMetadata(filename=filename or "document.docx")
    try:
        doc = Document(stream)
        parts = [p.text for p in doc.paragraphs if p.text]
        extracted = "\n".join(parts)
        cp = getattr(doc, "core_properties", None)
        if cp:
            metadata.author = getattr(cp, "author", None) or None
            metadata.modified_by = getattr(cp, "last_modified_by", None) or None
            mod = getattr(cp, "modified", None)
            metadata.modified = mod.isoformat() if mod else None
        normalized = _normalize(extracted)
        if not normalized:
            return ExtractedFileResult(
                extracted_text="",
                metadata=metadata,
                status=STATUS_EMPTY,
                is_scanned_review_required=False,
                mime_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            )
        if _is_sparse(normalized):
            return ExtractedFileResult(
                extracted_text=extracted,
                metadata=metadata,
                status=STATUS_SCANNED_REVIEW_REQUIRED,
                is_scanned_review_required=True,
                mime_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            )
        return ExtractedFileResult(
            extracted_text=extracted,
            metadata=metadata,
            status=STATUS_OK,
            is_scanned_review_required=False,
            mime_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )
    except Exception:
        return ExtractedFileResult(
            extracted_text="",
            metadata=metadata,
            status=STATUS_EMPTY,
            is_scanned_review_required=False,
            mime_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )


def extract_txt(stream: BinaryIO, filename: str) -> ExtractedFileResult:
    """Extract text from plain text file."""
    metadata = FileMetadata(filename=filename or "document.txt")
    try:
        raw = stream.read()
        if isinstance(raw, bytes):
            raw = raw.decode("utf-8", errors="replace")
        extracted = raw
    except Exception:
        extracted = ""
    normalized = _normalize(extracted)
    if not normalized:
        return ExtractedFileResult(
            extracted_text="",
            metadata=metadata,
            status=STATUS_EMPTY,
            is_scanned_review_required=False,
            mime_type="text/plain",
        )
    return ExtractedFileResult(
        extracted_text=extracted,
        metadata=metadata,
        status=STATUS_OK,
        is_scanned_review_required=False,
        mime_type="text/plain",
    )


def extract_csv(stream: BinaryIO, filename: str) -> ExtractedFileResult:
    """Extract text from CSV. Rows joined with newlines, cells with tab."""
    metadata = FileMetadata(filename=filename or "document.csv")
    try:
        raw = stream.read()
        if isinstance(raw, bytes):
            raw = raw.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(raw))
        rows = []
        for row in reader:
            rows.append("\t".join(str(c) for c in row))
        extracted = "\n".join(rows)
    except Exception:
        extracted = ""
    normalized = _normalize(extracted)
    if not normalized:
        return ExtractedFileResult(
            extracted_text="",
            metadata=metadata,
            status=STATUS_EMPTY,
            is_scanned_review_required=False,
            mime_type="text/csv",
        )
    return ExtractedFileResult(
        extracted_text=extracted,
        metadata=metadata,
        status=STATUS_OK,
        is_scanned_review_required=False,
        mime_type="text/csv",
    )


def extract_xlsx(stream: BinaryIO, filename: str) -> ExtractedFileResult:
    """Extract text from Excel XLSX. All sheets, cells joined as tab-separated rows."""
    from openpyxl import load_workbook

    metadata = FileMetadata(filename=filename or "document.xlsx")
    try:
        wb = load_workbook(stream, read_only=True, data_only=True)
        rows = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                rows.append("\t".join(cells))
        wb.close()
        extracted = "\n".join(rows)
    except Exception:
        extracted = ""
    normalized = _normalize(extracted)
    if not normalized:
        return ExtractedFileResult(
            extracted_text="",
            metadata=metadata,
            status=STATUS_EMPTY,
            is_scanned_review_required=False,
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    return ExtractedFileResult(
        extracted_text=extracted,
        metadata=metadata,
        status=STATUS_OK,
        is_scanned_review_required=False,
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def extract_xls(stream: BinaryIO, filename: str) -> ExtractedFileResult:
    """Extract text from legacy Excel XLS (97-2003). Requires xlrd."""
    import xlrd

    metadata = FileMetadata(filename=filename or "document.xls")
    try:
        wb = xlrd.open_workbook(file_contents=stream.read())
        rows = []
        for sheet in wb.sheets():
            for row_idx in range(sheet.nrows):
                cells = []
                for col_idx in range(sheet.ncols):
                    cell = sheet.cell(row_idx, col_idx)
                    val = cell.value
                    if val is not None:
                        cells.append(str(val))
                    else:
                        cells.append("")
                rows.append("\t".join(cells))
        extracted = "\n".join(rows)
    except Exception:
        extracted = ""
    normalized = _normalize(extracted)
    if not normalized:
        return ExtractedFileResult(
            extracted_text="",
            metadata=metadata,
            status=STATUS_EMPTY,
            is_scanned_review_required=False,
            mime_type="application/vnd.ms-excel",
        )
    return ExtractedFileResult(
        extracted_text=extracted,
        metadata=metadata,
        status=STATUS_OK,
        is_scanned_review_required=False,
        mime_type="application/vnd.ms-excel",
    )


def extract_file(file_like: BinaryIO, filename: str) -> ExtractedFileResult:
    """
    Single entry: raw file → extracted text + metadata.
    Raises ValueError for unsupported type.
    """
    if hasattr(file_like, "seek"):
        file_like.seek(0)
    fn = (filename or "").lower()
    if fn.endswith(".pdf"):
        return extract_pdf(file_like, filename or "document.pdf")
    if fn.endswith(".docx"):
        return extract_docx(file_like, filename or "document.docx")
    if fn.endswith(".txt"):
        return extract_txt(file_like, filename or "document.txt")
    if fn.endswith(".csv"):
        return extract_csv(file_like, filename or "document.csv")
    if fn.endswith(".xlsx"):
        return extract_xlsx(file_like, filename or "document.xlsx")
    if fn.endswith(".xls"):
        return extract_xls(file_like, filename or "document.xls")
    raise ValueError("Unsupported file type")
