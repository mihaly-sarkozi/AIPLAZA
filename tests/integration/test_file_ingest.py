# tests/integration/test_file_ingest.py
"""File ingest integration tests: plain text PDF, empty PDF, scanned fallback, DOCX metadata."""
from __future__ import annotations

import io

import pytest

from apps.knowledge.application.file_ingest import (
    extract_file,
    extract_txt,
    extract_docx,
    extract_pdf,
    ExtractedFileResult,
    FileMetadata,
    STATUS_OK,
    STATUS_EMPTY,
    STATUS_SCANNED_REVIEW_REQUIRED,
    MIN_EXTRACTED_TEXT_LEN,
)

pytestmark = pytest.mark.integration

# ---- Helpers: in-memory PDFs for tests ----
def _make_pdf_with_text(content: str) -> bytes:
    """Generate a one-page PDF with the given text (reportlab)."""
    pytest.importorskip("reportlab")
    from reportlab.pdfgen import canvas
    buf = io.BytesIO()
    c = canvas.Canvas(buf)
    c.drawString(100, 750, content[:200] if len(content) > 200 else content)
    c.save()
    return buf.getvalue()


def _make_empty_pdf() -> bytes:
    """Minimal valid PDF with one blank page (no extractable text)."""
    return (
        b"%PDF-1.4\n"
        b"1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj\n"
        b"2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj\n"
        b"3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] >> endobj\n"
        b"xref\n0 4\n0000000000 65535 f \n0000000009 00000 n \n0000000052 00000 n \n0000000101 00000 n \n"
        b"trailer << /Size 4 /Root 1 0 R >>\nstartxref\n178\n%%EOF"
    )


# ---- Plain text PDF ----
@pytest.mark.release_acceptance
def test_plain_text_pdf_extraction():
    """Plain text PDF: extracted text is non-empty, status ok, no scanned flag."""
    content = "Sample PDF content for ingest test. This has enough characters to pass the minimum length."
    pdf_bytes = _make_pdf_with_text(content)
    result = extract_file(io.BytesIO(pdf_bytes), "sample.pdf")
    assert result.status == STATUS_OK
    assert not result.is_scanned_review_required
    assert len(result.extracted_text.strip()) >= MIN_EXTRACTED_TEXT_LEN or content[:50] in result.extracted_text
    assert result.metadata.filename == "sample.pdf"


# ---- Empty PDF ----
@pytest.mark.release_acceptance
def test_empty_pdf_marked_as_empty_or_scanned():
    """Empty or no-text PDF: status empty or scanned_review_required, is_scanned_review_required True when sparse."""
    empty_pdf = _make_empty_pdf()
    result = extract_file(io.BytesIO(empty_pdf), "empty.pdf")
    assert result.status in (STATUS_EMPTY, STATUS_SCANNED_REVIEW_REQUIRED)
    assert result.is_scanned_review_required or (result.status == STATUS_EMPTY and not result.extracted_text.strip())
    assert result.metadata.filename == "empty.pdf"


# ---- Scanned PDF fallback (sparse text) ----
def test_scanned_pdf_fallback_sparse_text():
    """PDF with very little text (< MIN_EXTRACTED_TEXT_LEN): marked scanned_review_required."""
    short = "Only a few words."
    pdf_bytes = _make_pdf_with_text(short)
    result = extract_file(io.BytesIO(pdf_bytes), "scanned.pdf")
    assert result.status == STATUS_SCANNED_REVIEW_REQUIRED
    assert result.is_scanned_review_required
    assert result.extracted_text.strip() != ""


# ---- DOCX with metadata ----
@pytest.mark.release_acceptance
def test_docx_with_metadata():
    """DOCX: text extraction and author / last_modified_by in metadata."""
    from docx import Document
    from docx.opc.coreprops import CoreProperties
    buf = io.BytesIO()
    doc = Document()
    doc.add_paragraph("This is a test document with enough content to pass the minimum length for indexing.")
    doc.core_properties.author = "Test Author"
    doc.core_properties.last_modified_by = "Editor Name"
    doc.save(buf)
    buf.seek(0)
    result = extract_file(buf, "with_meta.docx")
    assert result.status == STATUS_OK
    assert "test document" in result.extracted_text.lower() or "enough content" in result.extracted_text.lower()
    assert result.metadata.filename == "with_meta.docx"
    assert result.metadata.author == "Test Author"
    assert result.metadata.modified_by == "Editor Name"


# ---- TXT ----
@pytest.mark.release_acceptance
def test_plain_txt_extraction():
    """Plain text file: content extracted, status ok."""
    text = "Plain text file content for ingest. Enough characters here."
    result = extract_txt(io.BytesIO(text.encode("utf-8")), "readme.txt")
    assert result.status == STATUS_OK
    assert result.extracted_text.strip() == text.strip()
    assert result.metadata.filename == "readme.txt"


def test_empty_txt():
    """Empty TXT: status empty."""
    result = extract_txt(io.BytesIO(b""), "empty.txt")
    assert result.status == STATUS_EMPTY
    assert result.extracted_text.strip() == ""


# ---- CSV ----
@pytest.mark.release_acceptance
def test_csv_extraction():
    """CSV: rows extracted as tab-separated, status ok."""
    csv_content = "name,email,phone\nJohn Doe,john@example.com,+36123456789\nJane,jane@test.hu,06201234567"
    result = extract_file(io.BytesIO(csv_content.encode("utf-8")), "data.csv")
    assert result.status == STATUS_OK
    assert "John Doe" in result.extracted_text
    assert "john@example.com" in result.extracted_text
    assert result.metadata.filename == "data.csv"


def test_empty_csv():
    """Empty CSV: status empty."""
    result = extract_file(io.BytesIO(b""), "empty.csv")
    assert result.status == STATUS_EMPTY


# ---- XLSX ----
@pytest.mark.release_acceptance
def test_xlsx_extraction():
    """XLSX: cell values extracted, status ok."""
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook
    buf = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Header1"
    ws["B1"] = "Header2"
    ws["A2"] = "Value one"
    ws["B2"] = "Value two with enough content for minimum length."
    wb.save(buf)
    buf.seek(0)
    result = extract_file(buf, "sheet.xlsx")
    assert result.status == STATUS_OK
    assert "Header1" in result.extracted_text
    assert "Value one" in result.extracted_text
    assert result.metadata.filename == "sheet.xlsx"


# ---- Unsupported type ----
def test_unsupported_file_type_raises():
    """Unsupported extension raises ValueError."""
    with pytest.raises(ValueError, match="Unsupported"):
        extract_file(io.BytesIO(b"x"), "file.xyz")


# ---- Service layer: train_from_file returns status empty / scanned_review_required ----
@pytest.mark.release_acceptance
def test_train_from_file_returns_empty_status_when_extract_empty():
    """When extract_file returns empty, train_from_file returns status 'empty' and does not call add_block."""
    import asyncio
    from unittest.mock import MagicMock, patch
    from apps.knowledge.application.knowledge_service import KnowledgeBaseService
    from apps.knowledge.application.file_ingest import (
        ExtractedFileResult,
        FileMetadata,
        STATUS_EMPTY,
    )

    mock_repo = MagicMock()
    mock_repo.get_by_uuid.return_value = MagicMock(id=1)
    mock_qdrant = MagicMock()
    svc = KnowledgeBaseService(repo=mock_repo, qdrant_service=mock_qdrant)

    empty_result = ExtractedFileResult(
        extracted_text="",
        metadata=FileMetadata(filename="empty.pdf"),
        status=STATUS_EMPTY,
        is_scanned_review_required=True,
    )
    fake_file = MagicMock()
    fake_file.filename = "empty.pdf"
    fake_file.file = io.BytesIO(b"")

    async def run():
        with patch("apps.knowledge.application.knowledge_service.extract_file", return_value=empty_result):
            return await svc.train_from_file("kb-uuid", fake_file, current_user_id=None)

    out = asyncio.run(run())
    assert out["status"] == "empty"
    assert "message" in out
    assert "metadata" in out
    mock_repo.add_training_log.assert_not_called()


@pytest.mark.release_acceptance
def test_train_from_file_returns_scanned_review_required_when_extract_sparse():
    """When extract_file returns scanned_review_required, train_from_file returns that status and does not add_block."""
    import asyncio
    from unittest.mock import MagicMock, patch
    from apps.knowledge.application.knowledge_service import KnowledgeBaseService
    from apps.knowledge.application.file_ingest import (
        ExtractedFileResult,
        FileMetadata,
        STATUS_SCANNED_REVIEW_REQUIRED,
    )

    mock_repo = MagicMock()
    mock_repo.get_by_uuid.return_value = MagicMock(id=1)
    mock_qdrant = MagicMock()
    svc = KnowledgeBaseService(repo=mock_repo, qdrant_service=mock_qdrant)

    sparse_result = ExtractedFileResult(
        extracted_text="Few words.",
        metadata=FileMetadata(filename="scanned.pdf"),
        status=STATUS_SCANNED_REVIEW_REQUIRED,
        is_scanned_review_required=True,
    )
    fake_file = MagicMock()
    fake_file.filename = "scanned.pdf"
    fake_file.file = io.BytesIO(b"")

    async def run():
        with patch("apps.knowledge.application.knowledge_service.extract_file", return_value=sparse_result):
            return await svc.train_from_file("kb-uuid", fake_file, current_user_id=None)

    out = asyncio.run(run())
    assert out["status"] == "scanned_review_required"
    assert "message" in out
    mock_repo.add_training_log.assert_not_called()
